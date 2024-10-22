import openpyxl
from deep_translator import GoogleTranslator
from tkinter import Tk, filedialog

# 初始化Tkinter
root = Tk()
root.withdraw()  # 隐藏主窗口

# 弹出文件选择对话框
file_path = filedialog.askopenfilename(title='选择一个Excel文件', filetypes=[('Excel文件', '*.xlsx')])

# 如果用户选择了文件
if file_path:
    # 加载Excel文件
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 初始化翻译器
    translator = GoogleTranslator(source='auto', target='zh-CN')

    # 遍历A列并将翻译结果写入B列
    for row in range(1, ws.max_row + 1):
        cell_value = ws[f'B{row}'].value
        if cell_value:  # 确保单元格不为空
            translated = translator.translate(cell_value)
            ws[f'C{row}'] = translated

    # 保存更改到原文件中
    wb.save(file_path)

    print("翻译完成！")
else:
    print("没有选择文件。")
